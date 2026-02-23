"""
books/views.py
Dooars Granthika — Books module views.

URL names used in templates (must match urls.py exactly):
  books:book_list        books:book_detail     books:book_create
  books:book_update      books:book_delete     books:stock_dashboard
  books:export_books     books:export_books_excel  books:update_stock

Context variables consumed by each template are noted above each view.
"""

import io
from datetime import date

from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView, DeleteView, DetailView, ListView, UpdateView,
)

from .forms import BookForm
from .models import Book, Category

# ── How many available copies counts as "low stock"
LOW_STOCK_THRESHOLD = 3


# ─────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────

def _filter_books(qs, request):
    """
    Apply q / category / stock GET params to a Book queryset.
    Used identically by the list view and both export views so
    the exported data always matches what the user sees on screen.
    """
    q        = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    stock    = request.GET.get("stock", "").strip()

    if q:
        qs = qs.filter(
            Q(title__icontains=q)    |
            Q(author__icontains=q)   |
            Q(isbn__icontains=q)     |
            Q(publisher__icontains=q)
        )
    if category:
        qs = qs.filter(category__slug=category)

    if stock == "available":
        qs = qs.filter(available_copies__gt=LOW_STOCK_THRESHOLD)
    elif stock == "low-stock":
        qs = qs.filter(
            available_copies__gt=0,
            available_copies__lte=LOW_STOCK_THRESHOLD,
        )
    elif stock == "out-stock":
        qs = qs.filter(available_copies=0)

    return qs


# ─────────────────────────────────────────────────────────────
# Book List
# book_list.html needs:
#   books, categories, page_obj, paginator
# ─────────────────────────────────────────────────────────────

def book_list(request):
    qs = Book.objects.select_related("category").all()
    qs = _filter_books(qs, request)

    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get("page"))

    return render(request, "books/book_list.html", {
        "books":      page_obj,          # page_obj IS iterable like a queryset
        "page_obj":   page_obj,
        "paginator":  paginator,
        "categories": Category.objects.all(),
    })


# ─────────────────────────────────────────────────────────────
# Book Detail
# book_detail.html needs:  book
# ─────────────────────────────────────────────────────────────

def book_detail(request, pk):
    book = get_object_or_404(Book.objects.select_related("category"), pk=pk)
    return render(request, "books/book_detail.html", {"book": book})


# ─────────────────────────────────────────────────────────────
# Book Create
# book_form.html needs:  form, categories
# ─────────────────────────────────────────────────────────────

def book_create(request):
    if request.method == "POST":
        form = BookForm(request.POST, request.FILES)
        if form.is_valid():
            book = form.save()
            messages.success(request, f"{book.title} was added successfully.")
            return redirect("books:book_detail", pk=book.pk)
    else:
        form = BookForm()

    return render(request, "books/book_form.html", {
        "form":       form,
        "categories": Category.objects.all(),
    })


# ─────────────────────────────────────────────────────────────
# Book Update
# book_form.html needs:  form, categories
# ─────────────────────────────────────────────────────────────

def book_update(request, pk):
    book = get_object_or_404(Book, pk=pk)
    if request.method == "POST":
        form = BookForm(request.POST, request.FILES, instance=book)
        if form.is_valid():
            form.save()
            messages.success(request, f"{book.title} was updated successfully.")
            return redirect("books:book_detail", pk=book.pk)
    else:
        form = BookForm(instance=book)

    return render(request, "books/book_form.html", {
        "form":       form,
        "categories": Category.objects.all(),
    })


# ─────────────────────────────────────────────────────────────
# Book Delete
# book_delete.html needs:  book
# ─────────────────────────────────────────────────────────────

def book_delete(request, pk):
    book = get_object_or_404(Book, pk=pk)
    if request.method == "POST":
        title = book.title
        book.delete()
        messages.success(request, f"{title} was deleted.")
        return redirect("books:book_list")

    return render(request, "books/book_delete.html", {"book": book})


# ─────────────────────────────────────────────────────────────
# Stock Dashboard
# book_stock_dashboard.html needs:
#   total_books, available_books, low_stock_count, out_of_stock_count
#   available_pct, low_pct, out_pct
#   category_stats  (list of dicts: name, total, available, pct)
#   most_issued     (books with .issue_count attribute)
#   recent_books
#   low_stock_list, low_threshold
# ─────────────────────────────────────────────────────────────

def stock_dashboard(request):
    all_books = Book.objects.select_related("category").all()

    total_books       = all_books.count()
    available_books   = all_books.filter(available_copies__gt=LOW_STOCK_THRESHOLD).count()
    low_stock_count   = all_books.filter(
                            available_copies__gt=0,
                            available_copies__lte=LOW_STOCK_THRESHOLD,
                        ).count()
    out_of_stock_count = all_books.filter(available_copies=0).count()

    def pct(n):
        return round(n / total_books * 100) if total_books else 0

    # Per-category stats for the progress bars
    category_stats = []
    for cat in Category.objects.all():
        cat_qs    = all_books.filter(category=cat)
        total     = cat_qs.count()
        available = cat_qs.filter(available_copies__gt=0).count()
        category_stats.append({
            "name":      cat.name,
            "total":     total,
            "available": available,
            "pct":       round(available / total * 100) if total else 0,
        })

    # Most-issued: books where issued_copies (total − available) is highest
    # Achieved by ordering ascending on available_copies (proxy for most issued)
    most_issued_qs = all_books.order_by("available_copies")[:5]
    most_issued = []
    for b in most_issued_qs:
        b.issue_count = b.issued_copies   # attach attribute expected by template
        most_issued.append(b)

    recent_books   = all_books.order_by("-created_at")[:5]
    low_stock_list = all_books.filter(
                         available_copies__gt=0,
                         available_copies__lte=LOW_STOCK_THRESHOLD,
                     ).order_by("available_copies")

    return render(request, "books/book_stock_dashboard.html", {
        "total_books":        total_books,
        "available_books":    available_books,
        "low_stock_count":    low_stock_count,
        "out_of_stock_count": out_of_stock_count,
        "available_pct":      pct(available_books),
        "low_pct":            pct(low_stock_count),
        "out_pct":            pct(out_of_stock_count),
        "category_stats":     category_stats,
        "most_issued":        most_issued,
        "recent_books":       recent_books,
        "low_stock_list":     low_stock_list,
        "low_threshold":      LOW_STOCK_THRESHOLD,
    })


# ─────────────────────────────────────────────────────────────
# Export preview page
# book_export.html needs:
#   preview_books, preview_more, total_count
#   total_copies_sum, available_copies_sum, issued_copies_sum
#   categories  (for filter strip)
# ─────────────────────────────────────────────────────────────

def export_books(request):
    qs = Book.objects.select_related("category").all()
    qs = _filter_books(qs, request)

    total_count   = qs.count()
    preview_limit = 8
    preview_books = list(qs[:preview_limit])
    preview_more  = max(0, total_count - preview_limit)

    agg = qs.aggregate(
        total_sum     = Sum("total_copies"),
        available_sum = Sum("available_copies"),
    )
    total_sum     = agg["total_sum"]     or 0
    available_sum = agg["available_sum"] or 0
    issued_sum    = total_sum - available_sum

    return render(request, "books/book_export.html", {
        "preview_books":       preview_books,
        "preview_more":        preview_more,
        "total_count":         total_count,
        "total_copies_sum":    total_sum,
        "available_copies_sum": available_sum,
        "issued_copies_sum":   issued_sum,
        "categories":          Category.objects.all(),
    })


# ─────────────────────────────────────────────────────────────
# Excel download
# ─────────────────────────────────────────────────────────────

def export_books_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(
            request,
            "openpyxl is not installed. Run: pip install openpyxl",
        )
        return redirect("books:export_books")

    qs = Book.objects.select_related("category").all()
    qs = _filter_books(qs, request)

    # ── Style constants ───────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="0A1628")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=9)
    TOTALS_FILL  = PatternFill("solid", fgColor="1E3A5F")
    TOTALS_FONT  = Font(color="FFFFFF", bold=True, size=9)
    GREEN_FILL   = PatternFill("solid", fgColor="DCFCE7")
    ORANGE_FILL  = PatternFill("solid", fgColor="FFF7ED")
    RED_FILL     = PatternFill("solid", fgColor="FEE2E2")
    CENTER       = Alignment(horizontal="center", vertical="center")
    THIN         = Side(style="thin", color="D1D5DB")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = openpyxl.Workbook()

    # ── Sheet 1 : Book Catalogue ──────────────────────────────
    ws = wb.active
    ws.title = "Book Catalogue"

    headers    = [
        "#", "Title", "Author", "ISBN", "Category", "Publisher",
        "Language", "Edition", "Total Copies", "Available", "Issued",
        "Shelf Location", "Added On",
    ]
    col_widths = [4, 32, 22, 18, 16, 20, 11, 12, 12, 10, 8, 14, 12]

    ws.append(headers)
    ws.row_dimensions[1].height = 22
    for ci, (_, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"

    for idx, book in enumerate(qs, 1):
        row_data = [
            idx,
            book.title,
            book.author,
            book.isbn,
            book.category.name if book.category else "",
            book.publisher,
            book.language,
            book.edition,
            book.total_copies,
            book.available_copies,
            book.issued_copies,
            book.shelf_location,
            book.created_at.strftime("%d/%m/%Y"),
        ]
        ws.append(row_data)
        r = ws.max_row

        # Colour-code the "Available" cell (column 10)
        avail_cell = ws.cell(row=r, column=10)
        if book.available_copies == 0:
            avail_cell.fill = RED_FILL
        elif book.available_copies <= LOW_STOCK_THRESHOLD:
            avail_cell.fill = ORANGE_FILL
        else:
            avail_cell.fill = GREEN_FILL

        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).border = BORDER

    last_data = ws.max_row

    # Totals row
    ws.append(
        ["", "TOTALS", "", "", "", "", "", "",
         f"=SUM(I2:I{last_data})",
         f"=SUM(J2:J{last_data})",
         f"=SUM(K2:K{last_data})",
         "", ""]
    )
    for ci in range(1, len(headers) + 1):
        cell           = ws.cell(row=ws.max_row, column=ci)
        cell.fill      = TOTALS_FILL
        cell.font      = TOTALS_FONT
        cell.alignment = CENTER
        cell.border    = BORDER

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_data}"

    # ── Sheet 2 : Stock Summary ───────────────────────────────
    ws2 = wb.create_sheet("Stock Summary")
    s2_headers = ["Category", "Total Books", "Available", "Issued", "Out of Stock"]
    ws2.append(s2_headers)
    ws2.row_dimensions[1].height = 22
    for ci, w in enumerate([22, 14, 14, 14, 14], 1):
        cell = ws2.cell(row=1, column=ci)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"

    for cat in Category.objects.all():
        cat_qs    = qs.filter(category=cat)
        total     = cat_qs.count()
        agg       = cat_qs.aggregate(
                        tc=Sum("total_copies"),
                        ac=Sum("available_copies"),
                    )
        tc        = agg["tc"] or 0
        ac        = agg["ac"] or 0
        available = cat_qs.filter(available_copies__gt=0).count()
        out       = cat_qs.filter(available_copies=0).count()
        ws2.append([cat.name, total, available, tc - ac, out])
        r = ws2.max_row
        for ci in range(1, 6):
            ws2.cell(row=r, column=ci).border = BORDER

    # ── Respond ───────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"dooars_granthika_books_{date.today():%Y%m%d}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ─────────────────────────────────────────────────────────────
# Stock bulk-upload (placeholder — wire up your upload template here)
# ─────────────────────────────────────────────────────────────

def update_stock(request):
    return render(request, "books/book_stock_update.html", {
        "categories": Category.objects.all(),
    })